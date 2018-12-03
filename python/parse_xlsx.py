#!/usr/bin/python

import openpyxl
import textwrap
import ast

baseline_db_xlsx = 'SX-9845-0182_Global_Extended_Metadata_Baseline_v1.14_2018-08-09.xlsx'

cell_version = 'C7'
cell_count = 'D7'
cell_table_start = 'C11'

sheet_teamids = 'TeamIDs'
cell_teamids_table_end = 'I492'

sheet_leagueids = 'LeagueIDs'
cell_leagueids_table_end = 'I17'

header_teamids = 'SxmTeamDb.h'
header_leagueids = 'SxmLeagueDb.h'


def makeTeamDbHeader():
    wb = openpyxl.load_workbook(baseline_db_xlsx)
    ws = wb.get_sheet_by_name(sheet_teamids)
    wb.close()

    hfile = open(header_teamids, 'w')

    hfile.write('#ifndef _' + header_teamids.replace('.', '_').upper() + '_\n')
    hfile.write('#define _' + header_teamids.replace('.', '_').upper() + '_\n')

    hfile.write('\n')
    hfile.write('#include <sxm_sdk.h>\n')

    hfile.write('\n')
    hfile.write('#define DB_TEAM_COUNT ' + str(ws[cell_count].value) + '\n')

    hfile.write('\n')
    hfile.write('const SXMTeam team_db[DB_TEAM_COUNT] = {\n')

    for r in ws[cell_table_start:cell_teamids_table_end]:
        id = r[0].value
        sn = r[1].value
        ln = r[2].value
        mn = r[3].value
        sports = r[4].value
        tiers = r[5].value

        id = int(id) if id is not None else -1
        sn = sn.encode('utf-8') if sn is not None else ''
        mn = mn.encode('utf-8') if mn is not None else ''
        ln = ln.encode('utf-8') if ln is not None else ''
        sports = sports.encode('utf-8') if sports is not None else ''
        tiers = tiers.encode('utf-8') if tiers is not None else ''
        cnt = len(ast.literal_eval(sports))
        
        sports = sports.replace('[', '{').replace(']', '}')
        tiers = tiers.replace('[', '{').replace(']', '}')

        hfile.write('\t{ ' + ', '.join([str(id), sn.join(['\"', '\"']), mn.join(['\"', '\"']), ln.join(['\"', '\"']), str(cnt), sports, tiers]) + '},\n')

    hfile.write('};\n')
    hfile.write('\n#endif')
    hfile.close()

def makeLeagueDbHeader():
    wb = openpyxl.load_workbook(baseline_db_xlsx)
    ws = wb.get_sheet_by_name(sheet_leagueids)
    wb.close()

    hfile = open(header_leagueids, 'w')

    hfile.write('#ifndef _' + header_leagueids.replace('.', '_').upper() + '_\n')
    hfile.write('#define _' + header_leagueids.replace('.', '_').upper() + '_\n')

    hfile.write('\n')
    hfile.write('#include <sxm_sdk.h>\n')

    hfile.write('\n')
    hfile.write('#define DB_LEAGUE_COUNT ' + str(ws[cell_count].value) + '\n')

    hfile.write('\n')
    hfile.write('const SXMLeague league_db[DB_LEAGUE_COUNT] = {\n')

    for r in ws[cell_table_start:cell_leagueids_table_end]:
        id = r[0].value
        sn = r[1].value
        ln = r[2].value
        lt = r[3].value
        sfe = r[4].value
        sft = r[5].value

        id = int(id) if id is not None else -1
        sn = sn.encode('utf-8') if sn is not None else ''
        ln = ln.encode('utf-8') if ln is not None else ''
        lt = lt.encode('utf-8') if lt is not None else ''
        sfe = int(sfe) if sfe is not None else -1
        sft = int(sft) if sft is not None else -1

        hfile.write('\t{ ' + ', '.join([str(id), sn.join(['\"', '\"']), ln.join(['\"', '\"']), lt.join(['\"', '\"']), str(sfe), str(sft)]) + '},\n')

    hfile.write('};\n')
    hfile.write('\n#endif')
    hfile.close()


if __name__ == '__main__':
    makeTeamDbHeader()
    makeLeagueDbHeader()
